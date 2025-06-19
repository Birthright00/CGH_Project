from openpyxl import load_workbook
import re
from datetime import datetime
import csv
import unicodedata

# Load workbook
wb = load_workbook("Copy of AY 2526 Phase III Medicine Posting Dates in Summary - For HCI Dated 21 Feb 25.xlsx")
ws = wb.active

# Collect all real date cells from the timetable (Mon–Fri)
date_cells = {}
for row in ws.iter_rows():
    for cell in row:
        if isinstance(cell.value, datetime):
            date_cells[cell.coordinate] = cell.value

# Flatten to datetime list
calendar_dates = list(date_cells.values())

# Month mapping for text matching
month_map = {
    'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
    'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
}

linked_results = []

# Helper: match a date range inside the calendar
def find_matching_dates_in_range(start_day, end_day, month, year_guess, remark_text, remark_cell):
    month_num = month_map[month]
    for dt in calendar_dates:
        if dt.year == year_guess and dt.month == month_num and start_day <= dt.day <= end_day:
            linked_results.append({
                "date": dt.strftime("%Y-%m-%d"),
                "remark": remark_text,
                "remark_cell": remark_cell
            })

# Go through all remark cells (column H)
# Go through all remark cells (column H)
for row in ws.iter_rows():
    cell = row[7]  # Column H
    if cell.value and isinstance(cell.value, str):
        # Normalize for safety (handles weird dash types, invisible chars)
        text = unicodedata.normalize("NFKC", cell.value.strip())

        # 🔧 Remove leading date portion like "27 Mar 26 Fri, " from remark
        cleaned_text = re.sub(
            r'^[\s,:–\-]*'                       # leading space/comma/colon/dash
            r'(?:'                               # non-capturing group for date/weekday
            r'\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)'
            r'(?:\s+\d{2,4})?(?:\s+\w{3})?'      # optional year and weekday
            r'|'
            r'(Mon|Tue|Wed|Thu|Fri|Sat|Sun)'     # OR weekday alone
            r')'
            r'[\s,:–\-]*',                       # trailing punctuation
            '',
            text,
            flags=re.UNICODE
        )


        if 'HOR Week' not in text and 'CBL Week' not in text:
            # --- Single date mentions: 27 Jun, 1 Aug
            single_dates = re.findall(r'\b(\d{1,2})\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\b', text)
            for day, month in single_dates:
                for dt in calendar_dates:
                    if dt.day == int(day) and dt.strftime('%b') == month:
                        linked_results.append({
                            "date": dt.strftime("%Y-%m-%d"),
                            "remark": cleaned_text,
                            "remark_cell": cell.coordinate
                        })

            # --- Full ranges: 9 Aug to 17 Aug 2025
            full_range = re.findall(r'(\d{1,2})\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+(?:to|-)\s+(\d{1,2})\s+\2(?:\s+(\d{4}))?', text)
            for start_day, month, end_day, year in full_range:
                year = int(year) if year else datetime.now().year
                find_matching_dates_in_range(int(start_day), int(end_day), month, year, cleaned_text, cell.coordinate)

            # --- Compact dash: 12–14 Sep or 1–3 Jul
            dash_range = re.findall(r'(\d{1,2})[\u2013-](\d{1,2})\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)', text)
            for start_day, end_day, month in dash_range:
                year_guess = max(dt.year for dt in calendar_dates if dt.strftime('%b') == month)
                find_matching_dates_in_range(int(start_day), int(end_day), month, year_guess, cleaned_text, cell.coordinate)

            # --- Compact single-month: 20 - 21 Oct
            compact_single_month_range = re.findall(r'(\d{1,2})\s*[\u2013\-]\s*(\d{1,2})\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)', text)
            for start_day, end_day, month in compact_single_month_range:
                year_guess = max(dt.year for dt in calendar_dates if dt.strftime('%b') == month)
                find_matching_dates_in_range(int(start_day), int(end_day), month, year_guess, cleaned_text, cell.coordinate)

        # --- HOR Week / CBL Week = Block entire row
        if 'HOR Week' in text or 'CBL Week' in text:
            for date_cell in row[:7]:  # Mon–Fri cols
                if isinstance(date_cell.value, datetime):
                    linked_results.append({
                        "date": date_cell.value.strftime("%Y-%m-%d"),
                        "remark": text,
                        "remark_cell": cell.coordinate
                    })

# --- De-duplicate results
seen = set()
deduped_results = []
for item in linked_results:
    key = (item['date'], item['remark_cell'])
    if key not in seen:
        seen.add(key)
        deduped_results.append(item)

# --- Sort
deduped_results.sort(key=lambda x: x['date'])

# --- Output
print("📌 Blocked Dates with Remarks:")
for item in deduped_results:
    print(f"{item['date']}: {item['remark']} (from {item['remark_cell']})")

# --- Get date-only list
blocked_dates = sorted(set(item['date'] for item in deduped_results))
print("\n🛑 Unique Blocked Dates:")
for d in blocked_dates:
    print(f"  {d}")

# --- Export to CSV
with open("blocked_dates_with_remarks.csv", "w", newline='') as f:
    writer = csv.DictWriter(f, fieldnames=["date", "remark", "remark_cell"])
    writer.writeheader()
    writer.writerows(deduped_results)

print("\n✅ Exported to blocked_dates_with_remarks.csv")
